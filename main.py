import os
import json
import shutil
import traceback
import re

from kivy.app import App
from kivy.lang import Builder
from kivy.utils import platform
from kivy.clock import Clock
from kivy.properties import StringProperty, BooleanProperty, NumericProperty, DictProperty, ListProperty
from kivy.metrics import dp
from kivy.uix.screenmanager import ScreenManager, Screen
from kivy.uix.recycleview import RecycleView
from kivy.uix.recycleboxlayout import RecycleBoxLayout
from kivy.uix.boxlayout import BoxLayout
from kivy.uix.label import Label
from kivy.uix.button import Button
from kivy.uix.popup import Popup
from kivy.uix.filechooser import FileChooserListView
from kivy.uix.scrollview import ScrollView
from kivy.uix.textinput import TextInput
from kivy.core.text import LabelBase

# --- 폰트 등록 ---
if os.path.exists("font.ttf"):
    try: LabelBase.register(name="Roboto", fn_regular="font.ttf")
    except: pass

# --- UI 디자인 ---
KV_UI = """
<Label>:
    font_name: 'Roboto'
<Button>:
    font_name: 'Roboto'
<TextInput>:
    font_name: 'Roboto'

<ListScreen>:
    BoxLayout:
        orientation: 'vertical'
        canvas.before:
            Color:
                rgba: 0.1, 0.1, 0.1, 1
            Rectangle:
                pos: self.pos
                size: self.size

        BoxLayout:
            size_hint_y: None
            height: '45dp'
            canvas.before:
                Color:
                    rgba: 0.15, 0.3, 0.4, 1
                Rectangle:
                    pos: self.pos
                    size: self.size
            Label:
                text: "현재 파일: " + app.current_filename
                bold: True
                size_hint_x: 0.7
            Button:
                text: "자동저장: " + ("ON" if app.auto_save else "OFF")
                size_hint_x: 0.3
                background_color: (0.2, 0.6, 0.2, 1) if app.auto_save else (0.6, 0.2, 0.2, 1)
                on_release: app.toggle_auto_save()

        BoxLayout:
            size_hint_y: None
            height: '60dp'
            padding: '5dp'
            spacing: '5dp'
            Button:
                text: '설정'
                on_release: app.open_smb_settings()
            Button:
                text: '엑셀선택'
                on_release: app.select_source('file')
            Button:
                text: 'PDF폴더'
                on_release: app.select_source('dir')
            Button:
                text: '리셋'
                background_color: 0.8, 0.3, 0.3, 1
                on_release: app.show_reset_confirm()
            Button:
                text: '저장'
                background_color: 0.2, 0.7, 0.3, 1
                on_release: app.save_to_excel()

        BoxLayout:
            size_hint_y: None
            height: '35dp'
            canvas.before:
                Color:
                    rgba: 0.2, 0.2, 0.2, 1
                Rectangle:
                    pos: self.pos
                    size: self.size
            Button:
                text: 'No' + app.sort_indicator_no
                size_hint_x: 0.08
                on_release: app.sort_by('no')
            Button:
                text: '품목코드' + app.sort_indicator_code
                size_hint_x: 0.24
                on_release: app.sort_by('item_code')
            Button:
                text: '수량' + app.sort_indicator_qty
                size_hint_x: 0.08
                on_release: app.sort_by('quantity')
            Button:
                text: '완료' + app.sort_indicator_comp
                size_hint_x: 0.13
                on_release: app.sort_by('complete')
            Button:
                text: '부족' + app.sort_indicator_short
                size_hint_x: 0.13
                on_release: app.sort_by('shortage')
            Button:
                text: '재작업' + app.sort_indicator_rew
                size_hint_x: 0.14
                on_release: app.sort_by('rework')
            Label:
                text: '비고'
                size_hint_x: 0.2

        CheckSheetRV:
            id: rv
            viewclass: 'RowWidget'
            RecycleBoxLayout:
                default_size: None, dp(30)
                default_size_hint: 1, None
                size_hint_y: None
                height: self.minimum_height
                orientation: 'vertical'
                spacing: '1dp'

<RowWidget>:
    orientation: 'horizontal'
    padding: [1, 1]
    canvas.before:
        Color:
            rgba: 0.2, 0.2, 0.2, 1
        Rectangle:
            pos: self.pos
            size: self.size
    Label:
        text: root.no
        size_hint_x: 0.08
    Button:
        text: root.item_code
        size_hint_x: 0.24
        font_size: '13sp'
        background_normal: ''
        background_color: 0.2, 0.4, 0.6, 1
        on_release: root.open_pdf_external()
    Label:
        text: root.quantity
        size_hint_x: 0.08
    Button:
        text: 'V' if root.complete else ''
        size_hint_x: 0.13
        background_normal: ''
        background_color: (0, 0.8, 0, 0.6) if root.complete else (0.3, 0.3, 0.3, 1)
        on_release: root.on_status('complete')
    Button:
        text: 'V' if root.shortage else ''
        size_hint_x: 0.13
        background_normal: ''
        background_color: (0.8, 0.8, 0, 0.6) if root.shortage else (0.3, 0.3, 0.3, 1)
        on_release: root.on_status('shortage')
    Button:
        text: 'V' if root.rework else ''
        size_hint_x: 0.14
        background_normal: ''
        background_color: (0.8, 0, 0, 0.6) if root.rework else (0.3, 0.3, 0.3, 1)
        on_release: root.on_status('rework')
    TextInput:
        text: root.remarks
        size_hint_x: 0.2
        multiline: False
        font_size: '12sp'
        on_text: root.on_remarks_change(self.text)
"""

# --- Java 콜백용 전역 함수 ---
def update_status_from_java(item_code, status):
    app = App.get_running_app()
    if not app: return
    
    # UI 스레드에서 데이터 업데이트 수행
    def update_data(dt):
        rv = app.root.get_screen('list').ids.rv
        found = False
        for d in rv.data:
            if d['item_code'] == item_code:
                if status == 'complete':
                    d['complete'] = not d['complete']
                    if d['complete']: d['shortage'] = d['rework'] = False
                elif status == 'shortage':
                    d['shortage'] = not d['shortage']
                    if d['shortage']: d['complete'] = d['rework'] = False
                elif status == 'rework':
                    d['rework'] = not d['rework']
                    if d['rework']: d['complete'] = d['shortage'] = False
                found = True; break
        
        if found:
            rv.refresh_from_data()
            if app.auto_save: app.save_to_excel(show_popup=False)
            
    Clock.schedule_once(update_data)

class ListScreen(Screen): pass
class CheckSheetRV(RecycleView): pass

class RowWidget(BoxLayout):
    no = StringProperty(''); item_code = StringProperty(''); quantity = StringProperty('')
    remarks = StringProperty(''); complete = BooleanProperty(False); shortage = BooleanProperty(False); rework = BooleanProperty(False)
    def on_status(self, st): App.get_running_app().update_item_status(self.item_code, self.no, st)
    def on_remarks_change(self, text): App.get_running_app().update_remarks_data(self.item_code, self.no, text)
    # 내부 뷰어로 실행하도록 변경
    def open_pdf_external(self): App.get_running_app().open_pdf_viewer_flow(self.item_code)

class CheckSheetApp(App):
    # ... (생략된 기존 속성들) ...

    # 내부 PDF 뷰어 실행 로직 (새로 추가)
    def open_pdf_viewer_flow(self, item_code):
        rv = self.root.get_screen('list').ids.rv.data
        if not rv: return

        # PDF 폴더 경로 확인
        base_path = self.pdf_folder_path if self.pdf_folder_path else self.LOCAL_BASE
        
        # 파일 목록 및 현재 인덱스 준비
        file_list = []
        target_index = 0
        
        # 전체 리스트를 순회하며 PDF 경로 목록 생성
        for i, d in enumerate(rv):
            ic = d['item_code']
            p = os.path.join(base_path, f"{ic}.pdf")
            # SMB 경로일 경우 임시 로컬 경로로 미리 변환 (필요시 다운로드는 Java에서 하거나 여기서 처리)
            # 일단 단순화를 위해 로컬 경로 위주로 리스트 생성
            file_list.append(p)
            if ic == item_code: target_index = i

        try:
            from jnius import autoclass
            ArrayList = autoclass('java.util.ArrayList')
            PdfActivity = autoclass('org.example.checksheetv163.PdfActivity')

            j_list = ArrayList()
            for f in file_list: j_list.add(f)

            # Java PdfActivity 실행
            PdfActivity.open(j_list, target_index)
        except Exception as e:
            self.show_popup("오류", f"내부 뷰어 실행 실패: {str(e)}\n빌드 설정을 확인하세요.")

    def ask_permissions(self):
        if platform != 'android': return
        try:
            from android.permissions import request_permissions, Permission
            from jnius import autoclass
            request_permissions([Permission.READ_EXTERNAL_STORAGE, Permission.WRITE_EXTERNAL_STORAGE, Permission.INTERNET])
            Env = autoclass('android.os.Environment')
            if hasattr(Env, 'isExternalStorageManager') and not Env.isExternalStorageManager():
                mActivity = autoclass('org.kivy.android.PythonActivity').mActivity
                Intent = autoclass('android.content.Intent'); Settings = autoclass('android.provider.Settings')
                uri = autoclass('android.net.Uri').fromParts("package", mActivity.getPackageName(), None)
                intent = Intent(Settings.ACTION_MANAGE_APP_ALL_FILES_ACCESS_PERMISSION); intent.setData(uri)
                mActivity.startActivity(intent)
        except: pass

    def open_pdf_in_external_app(self, item_code):
        base_path = self.pdf_folder_path if self.pdf_folder_path else self.LOCAL_BASE
        
        # SMB 경로 처리
        if base_path.startswith("smb://"):
            conn = self.get_smb_conn_only()
            if not conn: self.show_popup("알림", "SMB 연결 실패"); return
            try:
                # url 파싱: smb://ip/share/path/to/dir
                parts = base_path.replace("smb://", "").split("/", 1)
                if len(parts) < 2: self.show_popup("오류", "잘못된 SMB 경로"); return
                share_name = parts[1].split("/", 1)[0]
                inner_path = parts[1].split("/", 1)[1] if "/" in parts[1] else ""
                
                remote_file = os.path.join(inner_path, f"{item_code}.pdf").replace("\\", "/")
                local_file = os.path.join(self.LOCAL_BASE, f"{item_code}.pdf")
                
                # 스마트 다운로드: 파일이 존재하고 크기와 시간이 일치하면 다운로드 건너뜀
                should_download = True
                try:
                    attr = conn.getAttributes(share_name, remote_file)
                    if os.path.exists(local_file):
                        local_size = os.path.getsize(local_file)
                        local_mtime = os.path.getmtime(local_file)
                        
                        # 1. 크기가 같고 2. 서버 파일이 로컬 파일보다 최신이 아니면 (작거나 같으면)
                        # 원격 시간(attr.last_write_time)과 로컬 시간(local_mtime) 비교
                        if local_size == attr.file_size and attr.last_write_time <= local_mtime:
                            should_download = False
                except: pass # 오류 시 안전을 위해 다운로드 진행

                if should_download:
                    if not os.path.exists(self.LOCAL_BASE): os.makedirs(self.LOCAL_BASE)
                    with open(local_file, 'wb') as lf:
                        conn.retrieveFile(share_name, remote_file, lf)
                    # 다운로드 후 로컬 파일 시간을 원격 파일 시간과 맞추면 더 정확하지만,
                    # 현재 로직(원격 > 로컬일 때만 다운)으로도 충분히 목적 달성 가능합니다.
                pdf_path = local_file
            except Exception as e:
                self.show_popup("오류", f"SMB 파일 다운로드 실패:\n{str(e)}"); return
        else:
            pdf_path = os.path.join(base_path, f"{item_code}.pdf")

        if not os.path.exists(pdf_path):
            self.show_popup("알림", f"파일 없음: {item_code}.pdf"); return
            
        if platform == 'android':
            try:
                from jnius import autoclass, cast
                mActivity = autoclass('org.kivy.android.PythonActivity').mActivity
                Intent = autoclass('android.content.Intent'); Uri = autoclass('android.net.Uri'); File = autoclass('java.io.File')
                StrictMode = autoclass('android.os.StrictMode'); StrictMode.disableDeathOnFileUriExposure()
                file_obj = File(pdf_path); uri_obj = Uri.fromFile(file_obj)
                intent = Intent(Intent.ACTION_VIEW); intent.setDataAndType(uri_obj, "application/pdf")
                intent.addFlags(Intent.FLAG_ACTIVITY_NEW_TASK); intent.addFlags(Intent.FLAG_GRANT_READ_URI_PERMISSION)
                mActivity.startActivity(intent)
            except Exception as e: self.show_popup("오류", str(e))
        else: self.show_popup("알림", f"윈도우 경로: {pdf_path}")

    def load_excel_data(self, path):
        try:
            from openpyxl import load_workbook
            wb = load_workbook(path, data_only=True); ws = wb.active; rows = list(ws.rows)
            if not rows: return
            h = [str(c.value).strip().lower() if c.value else "" for c in rows[0]]
            def find_col(name_list, default_idx):
                for name in name_list:
                    if name in h: return h.index(name)
                return default_idx
            idx_no = find_col(['no', '번호'], 0); idx_code = find_col(['품목코드', 'code'], 1); idx_qty = find_col(['수량', 'qty'], 2)
            idx_comp = find_col(['완료', 'done'], 3); idx_short = find_col(['수량부족', 'short'], 4); idx_rew = find_col(['재작업', 'rework'], 5); idx_rem = find_col(['비고', 'remarks'], 6)
            rv_data = []
            for i, row in enumerate(rows[1:]):
                if len(row) <= idx_code or not row[idx_code].value: continue
                rv_data.append({
                    'no': str(row[idx_no].value or ''), 'item_code': str(row[idx_code].value or ''), 'quantity': str(row[idx_qty].value or ''),
                    'remarks': str(row[idx_rem].value or '') if len(row) > idx_rem else '',
                    'complete': str(row[idx_comp].value or '').upper() == 'V' if len(row) > idx_comp else False,
                    'shortage': str(row[idx_short].value or '').upper() == 'V' if len(row) > idx_short else False,
                    'rework': str(row[idx_rew].value or '').upper() == 'V' if len(row) > idx_rew else False,
                    'real_index': i
                })
            self.root.get_screen('list').ids.rv.data = rv_data
            self.current_filename = os.path.basename(path); self.excel_path = path
        except Exception as e: self.show_popup("로드 오류", str(e))

    def show_reset_confirm(self):
        content = BoxLayout(orientation='vertical', padding=20, spacing=20)
        content.add_widget(Label(text="모든 체크와 비고를 지우시겠습니까?"))
        btn_layout = BoxLayout(size_hint_y=None, height=dp(50), spacing=dp(10))
        pop = Popup(title="데이터 리셋", content=content, size_hint=(0.8, 0.4))
        btn_layout.add_widget(Button(text="아니오", on_release=pop.dismiss))
        btn_layout.add_widget(Button(text="예", on_release=lambda x: (self.reset_all_data(), pop.dismiss())))
        content.add_widget(btn_layout)
        pop.open()

    def reset_all_data(self):
        rv = self.root.get_screen('list').ids.rv
        if not rv.data: return
        new_data = []
        for d in rv.data:
            d['complete'] = False; d['shortage'] = False; d['rework'] = False; d['remarks'] = ""
            new_data.append(d)
        rv.data = new_data
        rv.refresh_from_data()
        if self.auto_save: self.save_to_excel()

    def sort_by(self, col):
        rv = self.root.get_screen('list').ids.rv
        if not rv.data: return
        new_s = 'desc' if self.sort_states.get(col) == 'asc' else 'asc'
        self.sort_states = {col: new_s}
        
        # 모든 인디케이터 초기화
        for k in ['no', 'code', 'qty', 'comp', 'short', 'rew']: 
            setattr(self, f'sort_indicator_{k}', '')
        
        # 현재 컬럼의 인디케이터 설정
        map_col = {'no':'no', 'item_code':'code', 'quantity':'qty', 'complete':'comp', 'shortage':'short', 'rework':'rew'}
        setattr(self, f'sort_indicator_{map_col[col]}', " ▲" if new_s == 'asc' else " ▼")
        
        def natural_sort_key(s):
            val = s.get(col, '')
            if isinstance(val, bool): return 1 if val else 0
            return [int(text) if text.isdigit() else text.lower() for text in re.split('([0-9]+)', str(val))]
            
        rv.data = sorted(rv.data, key=natural_sort_key, reverse=(new_s == 'desc')); rv.refresh_from_data()

    def update_item_status(self, ic, no, st):
        rv = self.root.get_screen('list').ids.rv
        for d in rv.data:
            if d['item_code'] == ic and d['no'] == no:
                if st == 'complete':
                    d['complete'] = not d['complete']
                    if d['complete']: d['shortage'] = d['rework'] = False
                elif st == 'shortage':
                    d['shortage'] = not d['shortage']
                    if d['shortage']: d['complete'] = d['rework'] = False
                elif st == 'rework':
                    d['rework'] = not d['rework']
                    if d['rework']: d['complete'] = d['shortage'] = False
                break
        rv.refresh_from_data()
        if self.auto_save: self.save_to_excel(show_popup=False)

    def update_remarks_data(self, ic, no, text):
        rv = self.root.get_screen('list').ids.rv
        for d in rv.data:
            if d['item_code'] == ic and d['no'] == no:
                d['remarks'] = text; break
        if self.auto_save and self.excel_path: self.save_to_excel(show_popup=False)

    def save_to_excel(self, show_popup=True):
        if not self.excel_path: return
        try:
            from openpyxl import load_workbook
            wb = load_workbook(self.excel_path); ws = wb.active; h = [str(c.value).strip() if c.value else "" for c in ws[1]]
            target_cols = ['완료', '수량부족', '재작업', '비고']
            cols = {}
            for col_name in target_cols:
                if col_name in h: cols[col_name] = h.index(col_name) + 1
                else:
                    new_idx = len(h) + 1; ws.cell(row=1, column=new_idx).value = col_name
                    cols[col_name] = new_idx; h.append(col_name)
            for d in self.root.get_screen('list').ids.rv.data:
                r = d['real_index'] + 2
                ws.cell(row=r, column=cols['완료']).value = 'V' if d.get('complete') else ''
                ws.cell(row=r, column=cols['수량부족']).value = 'V' if d.get('shortage') else ''
                ws.cell(row=r, column=cols['재작업']).value = 'V' if d.get('rework') else ''
                ws.cell(row=r, column=cols['비고']).value = d.get('remarks', '')
            wb.save(self.excel_path)
            if show_popup: self.show_popup("알림", "저장 완료")
        except Exception as e: 
            if show_popup: self.show_popup("저장 실패", str(e))

    def select_source(self, mode):
        content = BoxLayout(orientation='vertical', padding=20, spacing=20)
        pop = Popup(title="선택", content=content, size_hint=(0.8, 0.4))
        content.add_widget(Button(text="내 휴대폰", on_release=lambda x: (pop.dismiss(), self.open_local_browser(mode))))
        content.add_widget(Button(text="PC 공유폴더", on_release=lambda x: (pop.dismiss(), self.open_smb_shares_browser(mode))))
        pop.open()

    def open_local_browser(self, mode, path=None):
        if path is None:
            # 최근 경로 기억 로직
            start_path = ""
            if mode == 'file' and self.excel_path:
                d = os.path.dirname(self.excel_path)
                if os.path.exists(d): start_path = d
            elif mode == 'dir' and self.pdf_folder_path:
                if os.path.exists(self.pdf_folder_path) and not self.pdf_folder_path.startswith("smb://"): 
                    start_path = self.pdf_folder_path
            
            if not start_path:
                start_path = "/storage/emulated/0" if platform=='android' else os.getcwd()
            path = start_path

        content = BoxLayout(orientation='vertical')
        lb = BoxLayout(orientation='vertical', size_hint_y=None); lb.bind(minimum_height=lb.setter('height')); scroll = ScrollView(); scroll.add_widget(lb)
        content.add_widget(scroll)
        
        pop = Popup(title=f"탐색: {path}", content=content, size_hint=(0.9, 0.9))
        
        self.current_local_selection = None

        def confirm(x):
            if mode == 'dir':
                self.pdf_folder_path = path; self.save_settings(); pop.dismiss()
            elif mode == 'file':
                if self.current_local_selection:
                    self.excel_path = self.current_local_selection; self.load_excel_data(self.excel_path); self.save_settings(); pop.dismiss()
                else: self.show_popup("알림", "파일을 먼저 선택해 주세요.")

        content.add_widget(Button(text="선택 완료", size_hint_y=None, height=dp(60), on_release=confirm))

        # 상위 폴더로 이동 버튼
        parent_dir = os.path.dirname(path)
        if parent_dir and parent_dir != path:
            up_btn = Button(text=".. (상위 폴더로)", size_hint_y=None, height=dp(50), background_color=(0.3, 0.3, 0.3, 1))
            def go_up(x, p=parent_dir): pop.dismiss(); self.open_local_browser(mode, p)
            up_btn.bind(on_release=go_up); lb.add_widget(up_btn)

        try:
            # 파일 및 폴더 목록 가져오기
            items = sorted(os.listdir(path))
            valid_exts = ['.xlsx', '.xls'] if mode == 'file' else ['.pdf']
            
            for item in items:
                full_p = os.path.join(path, item)
                is_dir = os.path.isdir(full_p)
                
                if not is_dir:
                    ext = os.path.splitext(item)[1].lower()
                    if ext not in valid_exts: continue

                btn = Button(text=item + ("/" if is_dir else ""), size_hint_y=None, height=dp(50))
                if is_dir:
                    # 폴더는 싱글 클릭 시 즉시 진입
                    def go_dir(instance, p=full_p): pop.dismiss(); self.open_local_browser(mode, p)
                    btn.bind(on_release=go_dir)
                else:
                    # 파일은 클릭 시 선택 표시
                    def select_file(instance, p=full_p):
                        self.current_local_selection = p
                        for child in lb.children: child.background_color = (1, 1, 1, 1)
                        instance.background_color = (0.2, 0.4, 0.6, 1)
                    btn.bind(on_release=select_file)
                lb.add_widget(btn)
        except Exception as e:
            lb.add_widget(Label(text=f"접근 불가: {str(e)}", size_hint_y=None, height=dp(50)))
            
        pop.open()

    def open_smb_shares_browser(self, mode):
        conn = self.get_smb_conn_only()
        if not conn: self.show_popup("알림", "SMB 실패"); return
        content = BoxLayout(orientation='vertical'); lb = BoxLayout(orientation='vertical', size_hint_y=None); lb.bind(minimum_height=lb.setter('height')); scroll = ScrollView(); scroll.add_widget(lb)
        pop = Popup(title="공유폴더", content=content, size_hint=(0.9, 0.9))
        try:
            from smb.SMBConnection import SMBConnection
            for s in conn.listShares():
                if s.isSpecial or s.name.endswith('$'): continue
                b = Button(text=s.name, size_hint_y=None, height=80); b.bind(on_release=lambda x, n=s.name: self.open_smb_files_browser(conn, n, "/", mode, pop)); lb.add_widget(b)
        except: pass
        content.add_widget(scroll); pop.open()

    def open_smb_files_browser(self, conn, share, path, mode, parent):
        content = BoxLayout(orientation='vertical')
        lb = BoxLayout(orientation='vertical', size_hint_y=None); lb.bind(minimum_height=lb.setter('height')); scroll = ScrollView(); scroll.add_widget(lb)
        content.add_widget(scroll)
        
        pop = Popup(title=f"SMB: {share}{path}", content=content, size_hint=(0.9, 0.9))
        
        # 선택 상태 추적 (엑셀 파일 선택용)
        self.current_smb_selection = None

        def confirm(x):
            if mode == 'dir':
                # 폴더 선택 모드: 현재 경로 확정
                self.pdf_folder_path = f"smb://{self.smb_config['ip']}/{share}{path}"
                self.show_popup("알림", f"SMB PDF 경로 설정:\n{share}{path}")
                self.save_settings(); pop.dismiss(); parent.dismiss()
            elif mode == 'file':
                # 파일 선택 모드: 선택된 파일 다운로드 및 로드
                if self.current_smb_selection:
                    fname, fpath = self.current_smb_selection
                    local = os.path.join(self.LOCAL_BASE, fname)
                    if not os.path.exists(self.LOCAL_BASE): os.makedirs(self.LOCAL_BASE)
                    with open(local, 'wb') as lf: conn.retrieveFile(share, fpath, lf)
                    self.load_excel_data(local); self.save_settings(); pop.dismiss(); parent.dismiss()
                else:
                    self.show_popup("알림", "파일을 먼저 선택해 주세요.")

        # 하단에 '선택 완료' 버튼 추가 (형식 통일)
        content.add_widget(Button(text="선택 완료", size_hint_y=None, height=dp(60), on_release=confirm))

        def refresh(cp):
            lb.clear_widgets()
            valid_exts = ['.xlsx', '.xls'] if mode == 'file' else ['.pdf']
            for f in conn.listPath(share, cp):
                if f.filename in ['.', '..']: continue
                if not f.isDirectory:
                    ext = os.path.splitext(f.filename)[1].lower()
                    if ext not in valid_exts: continue

                btn = Button(text=f.filename + ("/" if f.isDirectory else ""), size_hint_y=None, height=dp(50))
                def click(instance, file=f):
                    np = os.path.join(path, file.filename).replace("\\", "/")
                    if file.isDirectory:
                        # 폴더 클릭 시 하위 폴더로 이동 (새 팝업)
                        pop.dismiss()
                        self.open_smb_files_browser(conn, share, np, mode, parent)
                    elif mode == 'file':
                        # 파일 클릭 시 선택 상태 업데이트 및 시각적 강조
                        self.current_smb_selection = (file.filename, np)
                        for child in lb.children: child.background_color = (1, 1, 1, 1)
                        instance.background_color = (0.2, 0.4, 0.6, 1) # 선택된 파일 강조
                btn.bind(on_release=click); lb.add_widget(btn)
        
        refresh(path); pop.open()

    def get_smb_conn_only(self):
        try:
            from smb.SMBConnection import SMBConnection
            c = self.smb_config; conn = SMBConnection(c['user'], c['pass'], "App", c['ip'], use_ntlm_v2=True, is_direct_tcp=True)
            if conn.connect(c['ip'], 445, timeout=3): return conn
        except: return None

    def open_smb_settings(self):
        content = BoxLayout(orientation='vertical', padding=10)
        ips = TextInput(text=self.smb_config['ip'], multiline=False); usr = TextInput(text=self.smb_config['user'], multiline=False); pas = TextInput(text=self.smb_config['pass'], password=True, multiline=False)
        content.add_widget(Label(text="IP")); content.add_widget(ips); content.add_widget(Label(text="ID")); content.add_widget(usr); content.add_widget(Label(text="PW")); content.add_widget(pas)
        pop = Popup(title="SMB 설정", content=content, size_hint=(0.8, 0.6))
        def save(x): self.smb_config = {'ip':ips.text,'user':usr.text,'pass':pas.text}; self.save_settings(); pop.dismiss()
        content.add_widget(Button(text="저장", on_release=save)); pop.open()

    def load_settings(self):
        if os.path.exists(self.SETTINGS_FILE):
            try:
                with open(self.SETTINGS_FILE, 'r') as f:
                    d = json.load(f)
                    self.excel_path = d.get('excel_path', '')
                    self.pdf_folder_path = d.get('pdf_folder_path', '')
                    self.smb_config = d.get('smb_config', {'ip':'','user':'','pass':''})
                    self.auto_save = d.get('auto_save', True)
            except: pass

    def save_settings(self):
        with open(self.SETTINGS_FILE, 'w') as f:
            json.dump({
                'excel_path': self.excel_path,
                'pdf_folder_path': self.pdf_folder_path,
                'smb_config': self.smb_config,
                'auto_save': self.auto_save
            }, f)

    def show_popup(self, title, msg): Popup(title=title, content=Label(text=str(msg)), size_hint=(0.8, 0.4)).open()

if __name__ == '__main__':
    CheckSheetApp().run()
